from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from backtest.kline import Kline
from backtest.monitor import Monitor
from draw.candlestick_drawer import CandlestickDrawer
from strategies.base import BaseStrategy
from strategies.hourly_template import SimplePinbarStrategy


@dataclass
class BacktestConfig:
    symbol: str = "BTCUSDT"
    interval: str = "1h"
    max_hold_bars: int = 4
    fee_rate: float = 0.0005
    stop_loss_pnl_pct: float = 10.0
    take_profit_pnl_pct: float = 25.0


class Backtester:
    def __init__(self, config: BacktestConfig, strategy: BaseStrategy):
        self.config = config
        self.strategy = strategy
        self.monitor = Monitor(symbol=config.symbol, interval=config.interval)

    def _validate(self):
        if self.config.interval != "1h":
            raise ValueError("This backtest is designed for 1h candles only.")
        if self.config.max_hold_bars <= 0:
            raise ValueError("max_hold_bars must be > 0")
        if self.config.stop_loss_pnl_pct <= 0 or self.config.take_profit_pnl_pct <= 0:
            raise ValueError("stop_loss_pnl_pct and take_profit_pnl_pct must be > 0")

    @staticmethod
    def _append_timestamp(save_path: str) -> Path:
        path = Path(save_path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return path.with_name(f"{path.stem}_{ts}{path.suffix}")

    def _build_tp_sl_prices(self, entry_price: float, side: str, leverage: float):
        tp_move = (self.config.take_profit_pnl_pct / 100.0) / leverage
        sl_move = (self.config.stop_loss_pnl_pct / 100.0) / leverage
        if side == "long":
            tp_price = entry_price * (1.0 + tp_move)
            sl_price = entry_price * (1.0 - sl_move)
        elif side == "short":
            tp_price = entry_price * (1.0 - tp_move)
            sl_price = entry_price * (1.0 + sl_move)
        else:
            raise ValueError(f"unsupported side: {side}")
        return tp_price, sl_price

    @staticmethod
    def _calc_unlevered_return(side: str, entry_price: float, exit_price: float) -> float:
        if side == "long":
            return (exit_price - entry_price) / entry_price
        if side == "short":
            return (entry_price - exit_price) / entry_price
        raise ValueError(f"unsupported side: {side}")

    def _to_klines(self, candles_df: pd.DataFrame, symbol: str) -> List[Kline]:
        klines: List[Kline] = []
        for _, row in candles_df.iterrows():
            ts = row["open_time"]
            timestamp = int(pd.Timestamp(ts).timestamp())
            klines.append(
                Kline(
                    symbol=symbol,
                    interval=self.config.interval,
                    timestamp=timestamp,
                    open_price=float(row["open"]),
                    high_price=float(row["high"]),
                    low_price=float(row["low"]),
                    close_price=float(row["close"]),
                    volume=float(row["volume"]),
                )
            )
        return klines

    def _build_entry_window_klines(self, all_klines: List[Kline], entry_idx: int) -> List[Kline]:
        start = max(0, entry_idx - 4)
        end = min(len(all_klines), entry_idx + 5)
        return all_klines[start:end]

    @staticmethod
    def _format_dt(value: Any) -> str:
        return pd.to_datetime(value).strftime("%Y-%m-%d %H:%M:%S")

    def run(
        self,
        start_time: Optional[str] = None,
        end_time: Optional[str] = None,
        limit: int = 1000,
    ) -> Dict[str, Any]:
        self._validate()
        candles_df = self.monitor.get_klines(
            symbol=self.config.symbol,
            interval=self.config.interval,
            start_time=start_time,
            end_time=end_time,
            limit=limit,
        ).reset_index(drop=True)
        klines = self._to_klines(candles_df, symbol=self.config.symbol)

        trades: List[Dict[str, Any]] = []
        position: Optional[Dict[str, Any]] = None

        for i in range(len(klines)):
            kline = klines[i]
            close_price = kline.close_price
            high_price = kline.high_price
            low_price = kline.low_price
            open_time = pd.to_datetime(kline.timestamp, unit="s")

            if position is None:
                signal = self.strategy.generate_entry_signal(i, klines)
                if signal:
                    side = signal["side"]
                    leverage = float(signal["leverage"])
                    tp_price, sl_price = self._build_tp_sl_prices(close_price, side, leverage)
                    position = {
                        "entry_idx": i,
                        "entry_time": open_time,
                        "side": side,
                        "leverage": leverage,
                        "signal": signal.get("signal"),
                        "amplitude_pct": signal.get("amplitude_pct"),
                        "entry_price": close_price,
                        "tp_price": tp_price,
                        "sl_price": sl_price,
                    }
                continue

            hold_bars = i - position["entry_idx"]
            strategy_exit = self.strategy.should_close(i, klines, position)
            timeout_exit = hold_bars >= self.config.max_hold_bars
            side = position["side"]
            tp_price = float(position["tp_price"])
            sl_price = float(position["sl_price"])

            if side == "long":
                hit_tp = high_price >= tp_price
                hit_sl = low_price <= sl_price
            else:
                hit_tp = low_price <= tp_price
                hit_sl = high_price >= sl_price

            exit_reason = None
            exit_price = close_price

            if hit_tp and hit_sl:
                # Conservative assumption for bar-level backtest ambiguity.
                exit_reason = "stop_loss_same_bar"
                exit_price = sl_price
            elif hit_sl:
                exit_reason = "stop_loss"
                exit_price = sl_price
            elif hit_tp:
                exit_reason = "take_profit"
                exit_price = tp_price
            elif strategy_exit:
                exit_reason = "strategy"
            elif timeout_exit:
                exit_reason = "max_hold_4h"

            if exit_reason is not None:
                entry = position["entry_price"]
                leverage = float(position["leverage"])
                gross_unlevered_return = self._calc_unlevered_return(side, entry, exit_price)
                gross_levered_return = gross_unlevered_return * leverage
                net_levered_return = gross_levered_return - 2 * self.config.fee_rate * leverage

                trades.append(
                    {
                        "entry_idx": position["entry_idx"],
                        "entry_time": position["entry_time"],
                        "exit_time": open_time,
                        "side": side,
                        "leverage": leverage,
                        "signal": position.get("signal"),
                        "amplitude_pct": position.get("amplitude_pct"),
                        "entry_price": entry,
                        "exit_price": exit_price,
                        "tp_price": tp_price,
                        "sl_price": sl_price,
                        "hold_bars": hold_bars,
                        "exit_reason": exit_reason,
                        "gross_unlevered_return": gross_unlevered_return,
                        "gross_levered_return": gross_levered_return,
                        "net_levered_return": net_levered_return,
                    }
                )
                position = None

        total_net_levered_return = sum(t["net_levered_return"] for t in trades)
        result = {
            "symbol": self.config.symbol,
            "interval": self.config.interval,
            "max_hold_bars": self.config.max_hold_bars,
            "stop_loss_pnl_pct": self.config.stop_loss_pnl_pct,
            "take_profit_pnl_pct": self.config.take_profit_pnl_pct,
            "bars": len(klines),
            "trades": len(trades),
            "total_net_levered_return": total_net_levered_return,
            "trade_list": trades,
            "all_klines": klines,
        }
        return result

    def save_trades_xlsx(self, result: Dict[str, Any], output_path: str) -> str:
        side_map = {"long": "做多", "short": "做空"}
        reason_map = {
            "stop_loss_same_bar": "同根K线先止损",
            "stop_loss": "止损",
            "take_profit": "止盈",
            "strategy": "策略平仓",
            "max_hold_4h": "超时平仓(4小时)",
        }

        headers = [
            "序号",
            "交易对",
            "周期",
            "信号时间",
            "平仓时间",
            "方向",
            "信号类型",
            "Pinbar振幅(%)",
            "杠杆倍数",
            "开仓价",
            "止盈价",
            "止损价",
            "平仓价",
            "持有K线数",
            "平仓原因",
            "毛收益率(未杠杆)",
            "毛收益率(杠杆后)",
            "净收益率(杠杆后)",
            "止损规则(账户%)",
            "止盈规则(账户%)",
            "最大持有小时",
            "买入点前后4h蜡烛图",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Backtest"

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        charts_dir = Path(output_path).parent / "charts"
        charts_dir.mkdir(parents=True, exist_ok=True)

        all_klines: List[Kline] = result.get("all_klines") or []
        drawer = CandlestickDrawer(symbol=result["symbol"], interval=result["interval"])

        image_col_idx = len(headers)
        image_col_letter = get_column_letter(image_col_idx)
        ws.column_dimensions[image_col_letter].width = 70

        for idx, trade in enumerate(result["trade_list"], start=1):
            row_idx = idx + 1
            row_values = [
                idx,
                result["symbol"],
                result["interval"],
                self._format_dt(trade["entry_time"]),
                self._format_dt(trade["exit_time"]),
                side_map.get(trade["side"], trade["side"]),
                trade.get("signal"),
                trade.get("amplitude_pct"),
                trade["leverage"],
                trade["entry_price"],
                trade["tp_price"],
                trade["sl_price"],
                trade["exit_price"],
                trade["hold_bars"],
                reason_map.get(trade["exit_reason"], trade["exit_reason"]),
                trade["gross_unlevered_return"],
                trade["gross_levered_return"],
                trade["net_levered_return"],
                result["stop_loss_pnl_pct"],
                result["take_profit_pnl_pct"],
                result["max_hold_bars"],
            ]

            for col_idx, val in enumerate(row_values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

            if all_klines:
                window_klines = self._build_entry_window_klines(all_klines, int(trade["entry_idx"]))
                if window_klines:
                    entry_time = pd.to_datetime(trade["entry_time"])
                    chart_base = charts_dir / (
                        f"{result['symbol']}_{result['interval']}_entry_{idx}_"
                        f"{entry_time.strftime('%Y%m%d_%H%M%S')}.png"
                    )
                    chart_path = drawer._plot_klines(
                        window_klines,
                        title=f"{result['symbol']} {result['interval']} Entry {idx} (+/-4h)",
                        save_path=str(chart_base),
                        show=False,
                    )
                    img = XLImage(chart_path)
                    img.width = 560
                    img.height = 280
                    ws.add_image(img, f"{image_col_letter}{row_idx}")
                    ws.row_dimensions[row_idx].height = 210

        for col_idx in range(1, image_col_idx):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 16

        out = self._append_timestamp(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        return str(out.resolve())


if __name__ == "__main__":
    config = BacktestConfig(symbol="BTCUSDT", interval="1h", max_hold_bars=4)
    strategy = SimplePinbarStrategy()
    backtester = Backtester(config=config, strategy=strategy)

    summary = backtester.run(
        start_time="2026-01-01 00:00:00",
        end_time="2026-01-24 00:00:00",
        limit=1000,
    )
    xlsx_path = backtester.save_trades_xlsx(summary, output_path="backtest/results/pinbar_backtest_report.xlsx")
    print(f"回测结果已保存: {xlsx_path}")
